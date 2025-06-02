# Writeten by Chun-Hsiang Chao
# Date:20250602
import datetime
import openpyxl 
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill, NamedStyle, GradientFill
from openpyxl.utils import FORMULAE
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference

table_name=["Test1","Test2","Test3"]
interger_i=0
workbook=openpyxl.Workbook() #apt install python3-openpyxl
workbook.template=True

sheet=workbook.worksheets[0]
sheet.page_setup_orientation=sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize=sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight=0
sheet.page_setup.fitToWidth=1

sheet.print_area="A1:I10"

sheet['A1']='Hello'
sheet['B1']='World'
sheet.title='Test1'
headings = ["NAME", "ID", "EMPLOYEE NAME", "NUMBER", "START TIME", "STOP TIME"]
sheet.append(headings)
for x in range(3,10):
	for y in range(1,10):
		interger_i+=1
		sheet.cell(row=x,column=y,value=interger_i)
sheet['C1']='=sum(A3:I3)'

sheet=workbook.create_sheet(table_name[1],1)
sheet['A1']='Hello'
sheet['B1']='World'
for x in range(3,10):
	for y in range(1,10):
		interger_i+=1
		sheet.cell(row=x,column=y,value=interger_i)

sheet.insert_cols(idx=2)
sheet.insert_cols(idx=4, amount=4)
sheet.delete_cols(idx=4, amount=3)
sheet.delete_cols(idx=6)

sheet.insert_rows(idx=5, amount=3)
sheet.delete_rows(idx=6, amount=2)


sheet=workbook.create_sheet('202505',2)
sheet['A1']='Hello'
sheet.cell(row=2,column=1).value=10

sheet=workbook.worksheets[2]
sheet['B1']='World'

# Create a few styles
bold_font = Font(name="Arial",bold=True)
big_red_text = Font(color="00FF0000", size=20)
center_aligned_text = Alignment(horizontal="center")
double_border_side = Side(border_style="double",color="ff0000")
thin_border_side=Side(border_style="thin",color="000000")
square_border = Border(top=double_border_side,right=thin_border_side,bottom=double_border_side,left=thin_border_side)

sheet=workbook.worksheets[0]


red_background = PatternFill(fgColor="00FF0000", fill_type="solid")
sheet.conditional_formatting.add("A3:I9",CellIsRule(operator="greaterThan", formula=["59"], fill=red_background))
blue_background = PatternFill(fgColor="000000FF", fill_type="solid")
sheet.conditional_formatting.add("A3:I9",CellIsRule(operator="lessThan", formula=["10"], fill=blue_background))
green_background = PatternFill(fgColor="0000FF00", fill_type="solid")
sheet.conditional_formatting.add("A3:I9",CellIsRule(operator="between", formula=["20","30"], fill=green_background))


sheet["A1"].font = bold_font
sheet["B3"].font = big_red_text
sheet["B3"].alignment = center_aligned_text
sheet["A1"].border = square_border
sheet["A1"].fill = PatternFill("solid",fgColor="00FF0000")
sheet["C1"].fill = GradientFill(stop=("000000","FFFFFF"))
sheet["E1"]=datetime.datetime(2025,5,27)
sheet["E1"].number_format="yyyy-mm-dd"
sheet["D1"]=datetime.datetime.now()
sheet["D1"].number_format="yyyy-mm-dd h:mm:ss"
sheet["F1"]=0.123456
sheet["F1"].number_format="0.00"
sheet["H1"]='=COUNTIF(H3:H9,">26")'
sheet["I1"]="=AVERAGE(I3:I9)"




#sheet.freeze_panes="C2"
#sheet.auto_filter.ref="A1:I9"


header = NamedStyle(name="header")
header.font = Font(bold=True)
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment(horizontal="center", vertical="center")

header_row = sheet[2]
for cell in header_row:
	cell.style = header

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=20)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
workbook.add_named_style(highlight)
sheet["I1"].style="highlight"

date_sheet=workbook["202505"]
operations_sheet = workbook.create_sheet("Operations")
workbook.remove(operations_sheet)
workbook.copy_worksheet(date_sheet)
workbook["202505 Copy"].title="202506"


rows = [
    ["Product", "Online", "Store"],
    [1, 30, 45],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 25],
    [6, 25, 35],
    [7, 20, 40],
]

for row in rows:
	sheet.append(row)

chart = BarChart()
data = Reference(worksheet=sheet,
min_row=10,
max_row=17,
min_col=2,
max_col=3)

chart.add_data(data, titles_from_data=True)
sheet.add_chart(chart, "E20")

workbook.save('test.xlsx')
workbook.close()

