# Writeten by Chun-Hsiang Chao
# Date:20251121
import datetime
import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill, NamedStyle, GradientFill
from openpyxl.utils import FORMULAE
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
#from xlcalculator import Evaluator
#pip install xlcalculator --break-system-packages

table_name=[
"202601","202601預算",
"202602","202602預算",
"202603","202603預算",
"202604","202604預算",
"202605","202605預算",
"202606","202606預算",
"202607","202607預算",
"202608","202608預算",
"202609","202609預算",
"202610","202610預算",
"202611","202611預算",
"202612","202612預算"]
#print(len(table_name))

workbook=openpyxl.load_workbook('my_financial_2025.xlsx',data_only=False)

sheet=workbook.worksheets[0]
sheet.page_setup_orientation=sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize=sheet.PAPERSIZE_A4
sheet.page_setup.fitToHeight=0
sheet.page_setup.fitToWidth=1

shlist=workbook.sheetnames
lpos=(len(shlist)-1)
#print(lpos)
#print(shlist)
sheets=workbook._sheets


for k in range(0,len(table_name)-1,2):
    print(k)
    source_sheet=workbook["yearMonth"]
    workbook.copy_worksheet(source_sheet)
    workbook["yearMonth Copy"].title=table_name[k]
    new_sheet=workbook[table_name[k]]
    sheets.pop(lpos+1)
    sheets.insert(5,new_sheet)
    lpos+=1
    
    source_sheet=workbook["yearMonth預算"]
    workbook.copy_worksheet(source_sheet)
    workbook["yearMonth預算 Copy"].title=table_name[k+1]
    new_sheet=workbook[table_name[k+1]]
    sheets.pop(lpos+1)
    sheets.insert(6,new_sheet)
    lpos+=1
    
    sheet=workbook[table_name[k+1]]
    sheet['B2']=table_name[k+1][4:6]
    
    string_to_replace = "yearMonth"
    replacement_string ="'"+table_name[k]+"'"
    
    amountOfRows = sheet.max_row
    amountOfColumns = sheet.max_column
    
    i = 0
    for r in range(1,sheet.max_row+1):
        for c in range(1,sheet.max_column+1):
            s = sheet.cell(r,c).value
            if s != None and string_to_replace in str(s):
                sheet.cell(r,c).value = s.replace(string_to_replace,replacement_string)
                sheet.cell(r,c).data_type='f'
                print("row {}	col {}	: {}".format(r,c,s))
                i += 1
    print(i)



delete_table_name=[
"202501","202501預算",
"202502","202502預算",
"202503","202503預算",
"202504","202504預算",
"202505","202505預算",
"202506","202506預算",
"202507","202507預算",
"202508","202508預算",
"202509","202509預算",
"202510","202510預算",
"202511","202511預算",
"202512","202512預算"]
#print(len(delete_table_name))


for k in range(0,len(delete_table_name),1):
    delete_sheet=workbook[delete_table_name[k]]
    workbook.remove(delete_sheet)




workbook.save('test_my_financial_2026.xlsx')
workbook.close()
