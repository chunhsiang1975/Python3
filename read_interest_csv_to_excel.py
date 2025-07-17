# Writeten by Chun-Hsiang Chao
# Date:20250716
import csv
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.chart.shapes import GraphicalProperties
from operator import itemgetter, attrgetter


def convertDate(date):
	str1=str(date)
	yearstr=str1[:3]
	realyear=str(int(yearstr)+1911)
	realdate=realyear+str1[4:6]+str1[7:9]
	return realdate


file_path = 'BWIBBU_d_ALL_20250701_utf-8.csv'


wb = Workbook()
sheet=wb.active
sheet.title = "CSV Data"  # Optional: Rename the sheet

with open(file_path, 'r', newline='\n', encoding='utf-8') as f:
	reader = csv.reader(f,delimiter=',')
	for row in reader:
		sheet.append(row)

records_number=len(sheet['A'])-2

int_columns=[6]
for x in range(3,records_number+2):
    for y in int_columns:
      if sheet.cell(row=x,column=y).value == "-" :
        sheet.cell(row=x,column=y,value="0")

int_columns=[3,6]
for x in range(3,records_number+2):
  for y in int_columns:
    s_list=list(sheet.cell(row=x,column=y).value)
    new_s_list=[]
    for z in range(0,len(s_list)):
      if (s_list[z]!=","):
          new_s_list.append(s_list[z])
    new_s=''.join(new_s_list)
    sheet.cell(row=x,column=y,value=new_s)


int_columns=[6,7]
for x in range(3,records_number+2):
  for y in int_columns:
    sheet.cell(row=x,column=y,value=float(sheet.cell(row=x,column=y).value))
#    sheet.cell(row=x,column=y,value=str(sheet.cell(row=x,column=y).value))




data=[]
for row in sheet.iter_rows(min_row=3, values_only=True):
    data.append(list(row))


#sorted_data = sorted(data, key=lambda x: (x[3],x[5],x[6]),reverse=True)
sorted_data = sorted(data, key=lambda x: (x[3]),reverse=True)
#sorted_data = sorted(data, key=itemgetter(3,5),reverse=True)
sheet.delete_rows(3, sheet.max_row)

for row_index, row_data in enumerate(sorted_data, start=3): # Start from row 2 for data
    for col_index, value in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=value)

#for x in range(3,records_number+3):
#  for y in range(4,9):
#    sheet.cell(row=x,column=y,value=float(sheet.cell(row=x,column=y).value))
#
#for x in range(3,records_number+3):
#	sheet.cell(row=x,column=1,value=(convertDate(sheet.cell(row=x,column=1).value)))



i=get_column_letter(1)
sheet.column_dimensions[i].width=10
i=get_column_letter(2)
sheet.column_dimensions[i].width=16
i=get_column_letter(3)
sheet.column_dimensions[i].width=16
i=get_column_letter(4)
sheet.column_dimensions[i].width=16
i=get_column_letter(5)
sheet.column_dimensions[i].width=16
i=get_column_letter(6)
sheet.column_dimensions[i].width=16
i=get_column_letter(7)
sheet.column_dimensions[i].width=16



wb.save('output_excel_stock_interest.xlsx')
wb.close()
