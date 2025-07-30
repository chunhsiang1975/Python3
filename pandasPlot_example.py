# Writeten by Chun-Hsiang Chao
# Date:20250730
import pandas as pd
import csv
import openpyxl
from openpyxl import Workbook


file_path = 'STOCK_DAY_1101_202506_utf-8.csv'
temp_file_path='test.xlsx'

wb = Workbook()
sheet=wb.active
sheet.title = "CSV Data"  # Optional: Rename the sheet

with open(file_path, 'r', newline='\n', encoding='utf-8') as f:
  reader = csv.reader(f,delimiter=',')
  for row in reader:
    sheet.append(row)

title=sheet.cell(row=1,column=1).value

rows=sheet.max_row
records_number=rows-7


sheet.delete_rows(idx=records_number+3,amount=5)
sheet.delete_rows(idx=1)

wb.save(temp_file_path)
wb.close()



#df=pd.read_csv('STOCK_DAY_1101_202506_data.csv')
#df=pd.read_excel('test.xlsx',sheet_name='CSV_data')
df=pd.read_excel(temp_file_path)
print(len(df.columns))
print(df.loc[[0,11]])
print(df.columns)


#pd.options.display.max_rows=9999
#print(df.to_string()) 
#print(df.loc[[0,11]])
#print(df.head())
#print(df.tail())
#print(df.info())
#print(df)
#print(len(df.columns))
#print(df.columns.tolist())
