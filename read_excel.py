# Writeten by Chun-Hsiang Chao
# Date:20250528
import openpyxl #apt install python3-openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

workbook=openpyxl.load_workbook('my_financial_3.xlsx')
sheet=workbook.worksheets[0]
print(sheet.title,sheet['F1'].value)
print(sheet.max_row,sheet.max_column)

sheet=workbook["202504"]

red_background = PatternFill(fgColor="00FF0000",fill_type="solid")
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ["$D1>200"]
sheet.conditional_formatting.add("A1:O100", rule)



workbook.save('test_read.xlsx')
workbook.close()
