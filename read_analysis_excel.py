# Writeten by Chun-Hsiang Chao
# Date:20250602
import openpyxl #apt install python3-openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.drawing.image import Image

workbook=openpyxl.load_workbook('my_financial_3.xlsx')
sheet=workbook.worksheets[0]
print(sheet.title,sheet['F1'].value)
print(sheet.max_row,sheet.max_column)

sheet=workbook["202504"]

red_background = PatternFill(fgColor="00FF0000",fill_type="solid")
diff_style = DifferentialStyle(fill=red_background)

rule = Rule(type="expression", dxf=diff_style)
rule.formula = ["$D1>300"]
sheet.conditional_formatting.add("C1:D200", rule)

ruleP = Rule(type="expression", dxf=diff_style)
ruleP.formula = ["$P1>2000"]
sheet.conditional_formatting.add("O1:P200", ruleP)

color_scale_rule = ColorScaleRule(
start_type='num',start_value=0,start_color='FF00FF00', 
mid_type='num',mid_value=500,mid_color='FFFFF000', 
end_type='num',end_value=1000,end_color='FFFF0000') 
sheet.conditional_formatting.add("P2:P200", color_scale_rule)

data_bar_rule=DataBarRule(
start_type="num",start_value=0,
end_type="num",end_value=500,
color="FF00FF00")
sheet.conditional_formatting.add("D2:D200", data_bar_rule)

logo = Image("example_1.png")
# A bit of resizing to not fill the whole spreadsheet with the logo
logo.height = 150
logo.width = 150

sheet.add_image(logo, "A1")

workbook.save('test_read.xlsx')
workbook.close()
