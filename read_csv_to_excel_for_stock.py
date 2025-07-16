# Writeten by Chun-Hsiang Chao
# Date:20250716
import csv
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.chart.shapes import GraphicalProperties

def convertDate(date):
	str1=str(date)
	yearstr=str1[:3]
	realyear=str(int(yearstr)+1911)
	realdate=realyear+str1[4:6]+str1[7:9]
	return realdate


file_path = 'STOCK_DAY_1101_202507_utf-8.csv'


wb = Workbook()
sheet=wb.active
sheet.title = "CSV Data"  # Optional: Rename the sheet

with open(file_path, 'r', newline='\n', encoding='utf-8') as f:
	reader = csv.reader(f,delimiter=',')
	for row in reader:
		sheet.append(row)

title=sheet.cell(row=1,column=1).value

rows=sheet.max_row
records_number=rows-7
#records_number=12

for x in range(3,records_number+3):
  for y in range(4,9):
    if(sheet.cell(row=x,column=y).value[0]=="X"):
      sheet.cell(x,y).value="0"
    sheet.cell(row=x,column=y,value=float(sheet.cell(row=x,column=y).value))

for x in range(3,records_number+3):
	sheet.cell(row=x,column=1,value=(convertDate(sheet.cell(row=x,column=1).value)))

int_columns=[2,3,9]
for x in range(3,records_number+3):
  for y in int_columns:
    s_list=list(sheet.cell(row=x,column=y).value)
    new_s_list=[]
    for z in range(0,len(s_list)):
      if (s_list[z]!=","):
          new_s_list.append(s_list[z])
    new_s=''.join(new_s_list)
    sheet.cell(row=x,column=y,value=int(new_s))




i=get_column_letter(1)
sheet.column_dimensions[i].width=10
i=get_column_letter(2)
sheet.column_dimensions[i].width=16
i=get_column_letter(3)
sheet.column_dimensions[i].width=16
i=get_column_letter(8)
sheet.column_dimensions[i].width=10


#for row in range(3,19):
#    sheet["{}{}".format("E", row)].number_format = 'General'
#    sheet["{}{}".format("F", row)].number_format = 'General'
#    sheet["{}{}".format("G", row)].number_format = 'General'


#print(sheet.cell(row=3,column=5).value)
sheet_chart=wb.create_sheet("Chart",1)
sheet_chart['G3']=title
sheet_chart.merge_cells('G3:K3')

chart=LineChart()
#Method 1
#data = Reference(sheet,min_row=2,max_row=21,min_col=4,max_col=7)
#chart.add_data(data, titles_from_data=True)

#Method 2
color_name=["F0F000","FF0000","00FF00","0000FF"]
for i in range(0,4):
	data = Reference(sheet,min_row=3,max_row=records_number+2,min_col=4+i,max_col=4+i)
	series=Series(data,title=sheet.cell(row=2,column=4+i).value)
	series.graphicalProperties.line.solidFill = color_name[i]
	chart.append(series)



#chart.y_axis.scaling.min = 24  # Set the minimum value for the y-axis
#chart.y_axis.scaling.max = 29 # Set the maximum value for the y-axis
date_x=Reference(sheet,min_row=3,max_row=23,min_col=1,max_col=1)
chart.set_categories(date_x)
sheet_chart.add_chart(chart, "A6")

chart=BarChart()
datas_columns=[2]
for i in datas_columns:
	data = Reference(sheet,min_row=3,max_row=records_number+2,min_col=i,max_col=i)
	series=Series(data,title=sheet.cell(row=2,column=i).value)
	series.graphicalProperties.line.solidFill = color_name[0]
	chart.append(series)

date_x=Reference(sheet,min_row=3,max_row=23,min_col=1,max_col=1)
chart.set_categories(date_x)
sheet_chart.add_chart(chart, "J6")


chart=BarChart()
datas_columns=[3]
for i in datas_columns:
	data = Reference(sheet,min_row=3,max_row=records_number+2,min_col=i,max_col=i)
	series=Series(data,title=sheet.cell(row=2,column=i).value)
	series.graphicalProperties.line.solidFill = color_name[1]
	chart.append(series)

date_x=Reference(sheet,min_row=3,max_row=23,min_col=1,max_col=1)
chart.set_categories(date_x)
sheet_chart.add_chart(chart, "A22")

chart=BarChart()
datas_columns=[9]
for i in datas_columns:
	data = Reference(sheet,min_row=3,max_row=records_number+2,min_col=i,max_col=i)
	series=Series(data,title=sheet.cell(row=2,column=i).value)
	series.graphicalProperties.line.solidFill = color_name[2]
	chart.append(series)

date_x=Reference(sheet,min_row=3,max_row=23,min_col=1,max_col=1)
chart.set_categories(date_x)
sheet_chart.add_chart(chart, "J22")

wb.save('output_excel_file.xlsx')
wb.close()
