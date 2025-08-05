# Writeten by Chun-Hsiang Chao
# Date:20250805
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
matplotlib.rc('font',family='Noto Serif JP')

import plotly.graph_objs as go
from plotly.graph_objs import Scatter, Layout
from plotly.offline import iplot, plot, init_notebook_mode


import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numpy as np

def convertDate(date):
  str1=str(date)
  yearstr=str1[:3]
  realyear=str(int(yearstr)+1911)
  realdate=realyear
  return realdate


file_path = 'FMSRFK_1101_2025_utf-8.csv'
temp_file_path='test.xlsx'

wb = Workbook()
sheet=wb.active

with open(file_path, 'r', newline='\n', encoding='utf-8') as f:
  reader = csv.reader(f,delimiter=',')
  for row in reader:
    sheet.append(row)

title=sheet.cell(row=1,column=1).value
s_list=list(title)
new_s_list=[]
for i in range(0,17):
  new_s_list.append(s_list[i])
new_s=''.join(new_s_list)
sheet.title = new_s

rows=sheet.max_row
records_number=rows-7



for x in range(3,records_number+3):
  sheet.cell(row=x,column=1,value=(convertDate(sheet.cell(row=x,column=1).value)))

int_columns=[6,7,8]
for x in range(3,records_number+3):
  for y in int_columns:
    s_list=list(sheet.cell(row=x,column=y).value)
    new_s_list=[]
    for z in range(0,len(s_list)):
      if (s_list[z]!=","):
          new_s_list.append(s_list[z])
    new_s=''.join(new_s_list)
    sheet.cell(row=x,column=y,value=int(new_s))


int_columns=[9]
for x in range(3,records_number+3):
  for y in int_columns:
    sheet.cell(row=x,column=y,value=float(sheet.cell(row=x,column=y).value))


for x in range(3,records_number+3):
  for y in range(3,5):
    if(sheet.cell(row=x,column=y).value[0]=="X"):
      sheet.cell(x,y).value="0"
#    sheet.cell(row=x,column=y,value=float(sheet.cell(row=x,column=y).value))
    sheet.cell(row=x,column=y,value=str(sheet.cell(row=x,column=y).value))

sheet.delete_rows(idx=records_number+3,amount=5)
sheet.delete_rows(idx=1)

i=get_column_letter(1)
sheet.column_dimensions[i].width=10
i=get_column_letter(6)
sheet.column_dimensions[i].width=16
i=get_column_letter(7)
sheet.column_dimensions[i].width=16
i=get_column_letter(8)
sheet.column_dimensions[i].width=16

wb.save(temp_file_path)
wb.close()


#df=pd.read_csv('STOCK_DAY_1101_202506_data.csv')
#df=pd.read_excel('test.xlsx',sheet_name='CSV_data')
df=pd.read_excel(temp_file_path)
#print(len(df.columns))
#print(df.loc[[0,11]])
#print(df.columns)

#pd.options.display.max_rows=9999
#print(df.to_string()) 
#print(df.loc[[0,11]])
#print(df.head())
#print(df.tail())
#print(df.info())
#print(df)
#print(len(df.columns))
#print(df.columns.tolist())


init_notebook_mode(connected=True)
#Method 1
data = [
    Scatter(x=df['月份'], y=df['加權(A/B)平均價'], name='加權(A/B)平均價',mode='lines+markers'),
    Scatter(x=df['月份'], y=df['最低價'], name='最低價',mode='lines+markers'),
    Scatter(x=df['月份'], y=df['最高價'], name='最高價',mode='lines+markers')
]

layout = go.Layout(
    title=title,
    xaxis=dict(title='月份'),
    yaxis=dict(title='價格',tickangle=45),
)
fig=go.Figure(data=data,layout=layout)
iplot(fig)
plot(fig,auto_open=True,filename='scatter_plot.html')


#Method 2
#x_data = [1, 2, 3, 4, 5]
x_data = df['月份']
#y_data = [2, 3, 1, 4, 2]
y_data = df['成交金額(A)']
bar_trace = go.Bar(
    x=x_data,
    y=y_data,
    name='成交金額(A)'
)
layout = go.Layout(
    title=title,
    xaxis=dict(title='月份'),
    yaxis=dict(title='成交金額(A)')
)
fig = go.Figure(data=[bar_trace], layout=layout)
iplot(fig)
plot(fig, auto_open=True, filename='bar_plot_1.html')

x_data = df['月份']
y_data = df['成交股數(B)']
bar_trace = go.Bar(
    x=x_data,
    y=y_data,
    name='成交股數'
)
layout = go.Layout(
    title=title,
    xaxis=dict(title='月份'),
    yaxis=dict(title='成交股數(B)')
)
fig = go.Figure(data=[bar_trace], layout=layout)
iplot(fig)
plot(fig, auto_open=True, filename='bar_plot_2.html')

x_data = df['月份']
y_data = df['成交筆數']
bar_trace = go.Bar(
    x=x_data,
    y=y_data,
    name='成交筆數'
)
layout = go.Layout(
    title=title,
    xaxis=dict(title='月份'),
    yaxis=dict(title='成交筆數')
)
fig = go.Figure(data=[bar_trace], layout=layout)
iplot(fig)
plot(fig, auto_open=True, filename='bar_plot_3.html')

#Method 3
#df['日期'] = pd.to_datetime(df['日期'],format='%Y-%m-%d')  #轉換日期欄位為日期格式
#print(df['日期'])
df.plot(kind='line', figsize=(12, 6), x='月份', y=['加權(A/B)平均價', '最低價', '最高價'])  #繪製統計圖
plt.xticks(rotation=45)
plt.savefig("dataframe.png")
plt.show()
 
#print(np.random.rand(10))
#print(np.random.rand(10)*10)
#print(np.random.randint(0,5,10))

