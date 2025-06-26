# Writeten by Chun-Hsiang Chao
# Date:20250626
import csv
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.chart.shapes import GraphicalProperties

def convertDate(date):
	str1=str(date)
	yearstr=str1[:3]
	realyear=str(int(yearstr)+1911)
	realdate=realyear+str1[4:6]+str1[7:9]
	return realdate



file_path = 'STOCK_DAY_1101_202506_utf-8.csv'



wb = Workbook()
sheet=wb.active
sheet.title = "CSV Data"  # Optional: Rename the sheet

with open(file_path, 'r', newline='\n', encoding='utf-8') as f:
	reader = csv.reader(f,delimiter=',')
	for row in reader:
		sheet.append(row)


for x in range(3,22):
  for y in range(4,9):
    sheet.cell(row=x,column=y,value=float(sheet.cell(row=x,column=y).value))


for x in range(3,22):
	sheet.cell(row=x,column=1,value=(convertDate(sheet.cell(row=x,column=1).value)))

i=get_column_letter(1)
sheet.column_dimensions[i].width=10
i=get_column_letter(2)
sheet.column_dimensions[i].width=16
i=get_column_letter(3)
sheet.column_dimensions[i].width=16
i=get_column_letter(8)
sheet.column_dimensions[i].width=10


#for row in range(3,19):
#    sheet["{}{}".format("E", row)].number_format = 'General'
#    sheet["{}{}".format("F", row)].number_format = 'General'
#    sheet["{}{}".format("G", row)].number_format = 'General'


#print(sheet.cell(row=3,column=5).value)


chart=LineChart()
#Method 1
#data = Reference(sheet,min_row=2,max_row=21,min_col=4,max_col=7)
#chart.add_data(data_1, titles_from_data=True)

#Method 2
color_name=["F0F000","FF0000","00FF00","0000FF"]
for i in range(0,4):
	data = Reference(sheet,min_row=3,max_row=21,min_col=4+i,max_col=4+i)
	series=Series(data,title=sheet.cell(row=2,column=4+i).value)
	series.graphicalProperties.line.solidFill = color_name[i]
	chart.append(series)






chart.y_axis.scaling.min = 24  # Set the minimum value for the y-axis
chart.y_axis.scaling.max = 29 # Set the maximum value for the y-axis



sheet.add_chart(chart, "J2")



wb.save('output_excel_file.xlsx')
wb.close()
